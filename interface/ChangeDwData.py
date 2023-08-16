import os
import openpyxl
import datetime
import tkinter as tk
from tkinter import messagebox
from tkinter import filedialog
import subprocess
import pandas as pd
from interface import LogDisplay
from setting import UserInformation


def ltrim_zero(number_str):
    while len(number_str) > 0 and number_str[0] == '0':
        number_str = number_str[1:]
    if number_str == '':
            return '0'
    return number_str

def number_extract(number_str):
    negtive = ''
    if number_str[0] == '-':
        negtive = '-'
        number_str = number_str[1:]
    elif number_str[0] == '+':
        number_str = number_str[1:]
    if number_str[-3:] == '000':
        number_str = number_str[:-3]
        while len(number_str) > 0 and number_str[0] == '0':
            number_str = number_str[1:]
        if number_str == '':
            return '0'
        return negtive + number_str
    return '[錯誤]: 右邊三位數非000'
    

def find_coordinate_of_text_in_sheet(text,work_sheet):
    result = []
    for column in work_sheet.iter_cols():
        for cell in column:
            if cell.value is not None and str(cell.value).find(text) == 0:
                result.append([cell.row, cell.column])
    if len(result) == 0:
        result = [[-1,-1]]
    return result
    
def get_keys_coordinate(work_sheet):
    return find_coordinate_of_text_in_sheet('key:',work_sheet)

def get_field_names_coordinate(work_sheet):
    return find_coordinate_of_text_in_sheet('Field Name',work_sheet)

def get_field_name_coordinate(work_sheet,key_row,key_col,search_range=3):
    for i in range(1,search_range+1):
        for j in range(1,search_range+1):
            if work_sheet.cell(key_row+i,key_col+j).value == 'Field Name':
                return key_row+i,key_col+j

def get_action_coordinate(work_sheet,key_row,key_col,search_range=2):
    for i in range(1,search_range+1):
        for j in range(1,search_range+1):
            if work_sheet.cell(key_row+i-3,key_col+j+2).value == 'Action:':
                return key_row+i-3,key_col+j+2

def convert_xls_to_xlsx(input_file, output_file):
   # 讀取 .xls 檔案
   xls_data = pd.read_excel(input_file, header=None, sheet_name=None)

   # 建立新的 .xlsx 檔案
   with pd.ExcelWriter(output_file, engine='xlsxwriter') as writer:
       # 將每個工作表寫入 .xlsx 檔案
       for sheet_name, df in xls_data.items():
            df.to_excel(writer, sheet_name=sheet_name, header=False, index=False)

def load_excel(path):
    if path[-5:].lower() == '.xlsx':
        return openpyxl.load_workbook(path)
    elif path[-4:].lower() == '.xls':
        new_excel_file = path[:-4]+'.xlsx'
        if os.path.isfile(new_excel_file):
            os.remove(new_excel_file)
        convert_xls_to_xlsx(path, new_excel_file)
        return openpyxl.load_workbook(new_excel_file)
    else:
        return

def write_file(path, content_list, mode='w'):
    f = open(path, mode)
    for line in content_list:
        f.write(line + '\n')
    f.close()

class ChangeDwData:
    def __init__(self, root):
        self.user = UserInformation.UserInformation()

        # 創建改檔頁面
        self.root = root
        # 畫面上半部
        self.upper_frame = tk.Frame(self.root)
        self.upper_frame.pack(fill=tk.BOTH, expand=True)

        # 畫面左上
        self.upper_left_frame = tk.Frame(self.upper_frame, pady=10, padx=10)
        self.upper_left_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        self.listbox = tk.Listbox(self.upper_left_frame)# 放入列表選擇框  
        self.listbox.insert(1, '使用excel產出改檔程式')    # 第一個選項
        self.listbox.insert(2, '使用ods改檔sql產出改檔程式')      # 第二個選項
        self.listbox.bind('<<ListboxSelect>>', self.on_select)
        self.listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, pady=3, padx=3)

        # 畫面中上
        self.upper_right_frame = tk.Frame(self.upper_frame, pady=10, padx=10)
        self.upper_right_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        self.excel_gen_code_view(self.upper_right_frame)

        # 畫面下半部
        self.lower_frame = tk.Frame(self.root, pady=10, padx=10) 
        self.log_display = LogDisplay.LogDisplay(self.lower_frame)
        self.lower_frame.pack(fill=tk.BOTH, expand=True)

        
        # gen code內容儲存於以下List當中
        self.output_result_cont = []
        self.output_test_command_cont = []
        self.output_prod_command_cont = []
        


    def on_select(self, evt):
        # Note here that Tkinter passes an event object to onselect()
        w = evt.widget
        index = int(w.curselection()[0])
        value = w.get(index)
        self.log_display.info("使用者操作：選擇 - " + value)
        if index == 0:
            self.clear_window(self.upper_right_frame)
            self.excel_gen_code_view(self.upper_right_frame)
        elif index == 1:
            #self.clear_window(self.upper_right_frame)
            self.log_display.info("此功能尚未開發完成")


    def clear_window(self, frame):
        # 清空畫面
        for widget in frame.winfo_children():
            if not isinstance(widget, tk.Menu):
                widget.destroy()


    def excel_gen_code_view(self, frame):
        self.label_file_path = tk.Label(frame, text="改檔Excel文件路徑:")
        self.label_file_path.grid(row=0, column=0, padx=10, pady=5)
        self.entry_file_path = tk.Entry(frame)
        self.entry_file_path.grid(row=0, column=1, padx=10, pady=5)
        self.button_get_excel = tk.Button(frame, text="開啟檔案", command=self.get_change_dw_data_excel)
        self.button_get_excel.grid(row=0, column=2, padx=10, pady=5)
        self.label_icontect_no = tk.Label(frame, text="IContect單號:")
        self.label_icontect_no.grid(row=1, column=0, padx=10, pady=5)
        self.entry_icontect_no = tk.Entry(frame)
        self.entry_icontect_no.grid(row=1, column=1, padx=10, pady=5)
        self.label_encrypt_acn = tk.Label(frame, text="去識別化後帳號:")
        self.label_encrypt_acn.grid(row=2, column=0, padx=10, pady=5)
        self.entry_encrypt_acn = tk.Entry(frame)
        self.entry_encrypt_acn.grid(row=2, column=1, padx=10, pady=5)
        self.label_data_ym = tk.Label(frame, text="資料年月(YYYYMM，未指定可免填):")
        self.label_data_ym.grid(row=3, column=0, padx=10, pady=5)
        self.entry_data_ym = tk.Entry(frame)
        self.entry_data_ym.grid(row=3, column=1, padx=10, pady=5)
        self.label_attachment_no = tk.Label(frame, text="電聯單附件編號(產出結果檔名使用):")
        self.label_attachment_no.grid(row=4, column=0, padx=10, pady=5)
        self.entry_attachment_no = tk.Entry(frame)
        self.entry_attachment_no.grid(row=4, column=1, padx=10, pady=5)
        self.label_comment = tk.Label(frame, text="修改內容描述(產出結果檔註解使用):")
        self.label_comment.grid(row=5, column=0, padx=10, pady=5)
        self.entry_comment = tk.Entry(frame)
        self.entry_comment.grid(row=5, column=1, padx=10, pady=5)
        self.button_gen_code = tk.Button(frame, text="產出改檔程式", command=self.gen_code)
        self.button_gen_code.grid(row=6, column=1, columnspan=1, padx=10, pady=5)

    def get_change_dw_data_excel(self):
        # 選擇檔案後回傳檔案路徑與名稱
        file_path = filedialog.askopenfilename()
        self.entry_file_path.delete(0,tk.END)
        self.entry_file_path.insert(0,file_path)
        
        (input_folder, source_excel) = os.path.split(file_path)
        icontect_no = source_excel.split('_')[0] if '_' in source_excel else source_excel[:14]
        self.entry_icontect_no.delete(0,tk.END)
        self.entry_icontect_no.insert(0,icontect_no)


    def gen_code(self):
        # 取得使用者輸入資訊
        (input_folder, source_excel) = os.path.split(self.entry_file_path.get())
        icontect_no = self.entry_icontect_no.get()
        self.encrypt_acn = self.entry_encrypt_acn.get()
        data_ym = self.entry_data_ym.get()
        attachment_no = self.entry_attachment_no.get()
        comment = self.entry_comment.get()
        
        self.icontect_no = icontect_no
        self.data_ym = data_ym
        
        # 建立產出code的資料夾
        output_path = input_folder.replace('/','\\') + r'\產出結果輸出區'
        if not os.path.isdir(output_path):
            os.mkdir(output_path)
            
        # 產出結果檔名
        if icontect_no is not None and len(icontect_no) == 14:
            output_result_fp = output_path + '\\' + icontect_no + '_teradata_' + attachment_no + '.sql'
        else:
            self.log_display.error('IConnect單號格式錯誤，應為14碼數字!')

        # 產出結果檔表頭
        self.output_result_cont.append('/*******************************************************************************')
        self.output_result_cont.append('* 修改日期     : ' + str(datetime.datetime.now().date()))
        self.output_result_cont.append('* AUTHOR ID    : ' + self.user.get_user_bank_no() + ' ' + self.user.get_user_full_name())
        self.output_result_cont.append('* 聯繫單號     : ' + icontect_no)
        self.output_result_cont.append('* 修改內容描述 : ' + comment)
        self.output_result_cont.append('*******************************************************************************/')
        self.output_result_cont.append('')
        self.output_result_cont.append('SET WIDTH 300')
        self.output_result_cont.append('')
        self.output_result_cont.append('/* ACCT_NBR轉碼: ' + self.encrypt_acn + ' */')
        self.output_result_cont.append('')
        
        # 測試檔表頭
        self.output_test_command_cont.append('/*******************************************************************************')
        self.output_test_command_cont.append('* ' + icontect_no + ' 測試使用指令')
        self.output_test_command_cont.append('*******************************************************************************/')
        self.output_test_command_cont.append('')
        self.output_test_command_cont.append('/* ACCT_NBR轉碼: ' + self.encrypt_acn + ' */')
        self.output_test_command_cont.append('')
        
        # 查看營運資料表頭
        self.output_prod_command_cont.append('/*******************************************************************************')
        self.output_prod_command_cont.append('* ' + icontect_no + ' 查看營運資料指令')
        self.output_prod_command_cont.append('*******************************************************************************/')
        self.output_prod_command_cont.append('')
        self.output_prod_command_cont.append('/* ACCT_NBR轉碼: ' + self.encrypt_acn + ' */')
        self.output_prod_command_cont.append('')
        
        # 開啟 Excel 檔案
        wb = load_excel(input_folder + '\\' + source_excel)
        if wb is None:
            self.log_display.error('來源Excel讀取失敗')
            return
        
        # 讀取 Excel 裡所有工作表名稱
        wb_sheet_names = wb.sheetnames
        for wb_sheet_name in wb_sheet_names:
            if 'RPYC' in wb_sheet_name.upper():
                self.gen_rpyc(wb[wb_sheet_name])
            elif 'BOCT' in wb_sheet_name.upper() and 'BOCT_Q' not in wb_sheet_name.upper() and 'BOCT-Q' not in wb_sheet_name.upper():
                self.gen_boct(wb[wb_sheet_name])
            elif 'BOCT_Q' in wb_sheet_name.upper() or 'BOCT-Q' in wb_sheet_name.upper():
                self.gen_boct_q(wb[wb_sheet_name])
        write_file(output_result_fp, self.output_result_cont, mode='w')
        self.output_result_cont = []
        write_file(output_path + r'\查看營運資料指令.sql', self.output_prod_command_cont, mode='w')
        self.output_prod_command_cont = []
        write_file(output_path + r'\測試使用指令.sql', self.output_test_command_cont, mode='w')
        self.output_test_command_cont = []
        print(output_path)
        subprocess.Popen('explorer ' + output_path)
        self.log_display.info('產出改檔程式成功!')

    def gen_rpyc(self,work_sheet):
        keys_coordinate = get_keys_coordinate(work_sheet)
        for row, col in keys_coordinate:
            action_row, action_col = get_action_coordinate(work_sheet, row, col, search_range=2)
            action = work_sheet.cell(action_row,action_col+1).value.upper()
            acct_acn = work_sheet.cell(row,col+1).value.strip()
            journal_nbr = ltrim_zero(str(work_sheet.cell(row,col+3).value).strip())
            term_nbr = work_sheet.cell(row,col+4).value.strip()
            bucket_seq_nbr = ltrim_zero(str(work_sheet.cell(row,col+5).value).strip())
            print(action,acct_acn,journal_nbr,term_nbr,bucket_seq_nbr)
            
            if action == 'A(AMEND)' and acct_acn is not None:
                self.gen_rpyc_amend_cont(work_sheet,self.icontect_no,row,col,acct_acn,journal_nbr,term_nbr,bucket_seq_nbr)
            elif action == 'D(DELETE)' and acct_acn is not None:
                self.gen_rpyc_delete_cont(work_sheet,self.icontect_no,row,col,acct_acn,journal_nbr,term_nbr,bucket_seq_nbr)
            elif action not in ['A(AMEND)','D(DELETE)']:
                self.log_display.error('改檔EXCEL RPYC工作表Action非"A(AMEND)"或"D(DELETE)"，請修正!')
                return
            elif action == 'A(AMEND)' and acct_acn is None:
                self.output_result_cont.append('/* 無RPYC更改需求 */\n')
            elif action == 'D(DELETE)' and acct_acn is None:
                self.output_result_cont.append('/* 無RPYC刪除需求 */\n')

    def gen_rpyc_amend_cont(self,work_sheet,icontect_no,key_row,key_col,acct_acn,journal_nbr,term_nbr,bucket_seq_nbr):
        field_name_row, field_name_col = get_field_name_coordinate(work_sheet,key_row,key_col,search_range=3)
        if field_name_row is None or field_name_col is None:
            self.log_display.error('改檔EXCEL RPYC工作表"key:"的右下格子非"Field Name"，請修正!')
            return
        else:
            field_dict = {'COLLECTED-AMT':'COLLCTD_AMT','REPAY-AMT':'REPAY_AMT','ADJUSTED-AMT':'ADJSTD_AMT','TXN-AMT':'BUCKET_AMT','BASED-AMT':'BASD_AMT'}
            i = 1
            amend_list = []
            while work_sheet.cell(field_name_row+i,field_name_col).value is not None:
                field_name = work_sheet.cell(field_name_row+i,field_name_col).value.strip()
                occurs = work_sheet.cell(field_name_row+i,field_name_col+1).value.strip()
                before_amend = number_extract(work_sheet.cell(field_name_row+i,field_name_col+2).value.strip())
                after_amend = number_extract(work_sheet.cell(field_name_row+i,field_name_col+3).value.strip())
                amend_col = field_dict[field_name] + '_' + occurs
                amend_list.append([amend_col,before_amend,after_amend])
                i += 1
            # 上線程式
            self.output_result_cont.append('/*** RPYC(AMEND)部份 ***/')
            self.output_result_cont.append('-- EVENT_BKL_REPAY_TXN 台幣放款還款紀錄檔')
            amend_cols = ''
            for amend_col,before_amend,after_amend in amend_list:
                amend_cols += amend_col + ','
            self.output_result_cont.append('-- Col: ' + amend_cols)
            self.output_result_cont.append('-------------------------------------------------------------------------------------------')
            self.output_result_cont.append("SELECT '" + icontect_no + "'||DATE||' '||TIME||' BEFORE-VALUE=>' SEQ")
            self.output_result_cont.append('      ,INSTITUTE_ID, ACCT_NBR, TXN_DATE, JOURNAL_NBR, TERM_NBR, BUCKET_SEQ_NBR')
            for amend_col,before_amend,after_amend in amend_list:
                self.output_result_cont.append('      ,' + amend_col)
            self.output_result_cont.append('FROM DP_MCIF.EVENT_BKL_REPAY_TXN')
            self.output_result_cont.append("WHERE ACCT_NBR = '" + acct_acn + "'")
            self.output_result_cont.append("AND JOURNAL_NBR = '" + journal_nbr + "'")
            self.output_result_cont.append("AND TERM_NBR ='" + term_nbr + "'")
            self.output_result_cont.append("AND BUCKET_SEQ_NBR = '" + bucket_seq_nbr + "';")
            self.output_result_cont.append('')
            self.output_result_cont.append('UPDATE DP_MCIF.EVENT_BKL_REPAY_TXN')
            self.output_result_cont.append('SET')
            for i in range(len(amend_list)):
                comma = ' ' if i == 0 else ','
                amend_col = amend_list[i][0]
                after_amend = amend_list[i][2]
                self.output_result_cont.append('    ' + comma + amend_col + ' = ' + after_amend)
            self.output_result_cont.append("WHERE ACCT_NBR = '" + acct_acn + "'")
            self.output_result_cont.append("AND JOURNAL_NBR = '" + journal_nbr + "'")
            self.output_result_cont.append("AND TERM_NBR ='" + term_nbr + "'")
            self.output_result_cont.append("AND BUCKET_SEQ_NBR = '" + bucket_seq_nbr + "';")
            self.output_result_cont.append('')
            self.output_result_cont.append("SELECT '" + icontect_no + "'||DATE||' '||TIME||' AFTER-VALUE=>' SEQ")
            self.output_result_cont.append('      , INSTITUTE_ID, ACCT_NBR, TXN_DATE, JOURNAL_NBR, TERM_NBR, BUCKET_SEQ_NBR')
            for amend_col,before_amend,after_amend in amend_list:
                self.output_result_cont.append('      , ' + amend_col)
            self.output_result_cont.append('FROM DP_MCIF.EVENT_BKL_REPAY_TXN')
            self.output_result_cont.append("WHERE ACCT_NBR = '" + acct_acn + "'")
            self.output_result_cont.append("AND JOURNAL_NBR = '" + journal_nbr + "'")
            self.output_result_cont.append("AND TERM_NBR ='" + term_nbr + "'")
            self.output_result_cont.append("AND BUCKET_SEQ_NBR = '" + bucket_seq_nbr + "';")
            self.output_result_cont.append('')
            
            # 查看營運資料指令
            self.output_prod_command_cont.append('/*** RPYC(AMEND)部份 ***/')
            self.output_prod_command_cont.append('-- EVENT_BKL_REPAY_TXN 台幣放款還款紀錄檔')
            self.output_prod_command_cont.append('')
            self.output_prod_command_cont.append('-- 查看營運資料')
            amend_cols = ''
            for amend_col,before_amend,after_amend in amend_list:
                amend_cols += amend_col + ','
            self.output_prod_command_cont.append('-- Col: ' + amend_cols)
            self.output_prod_command_cont.append('-------------------------------------------------------------------------------------------')
            self.output_prod_command_cont.append("SELECT '" + icontect_no + "'||DATE||' '||TIME||' BEFORE-VALUE=>' SEQ")
            self.output_prod_command_cont.append('      ,INSTITUTE_ID, ACCT_NBR, TXN_DATE, JOURNAL_NBR, TERM_NBR, BUCKET_SEQ_NBR')
            for amend_col,before_amend,after_amend in amend_list:
                self.output_prod_command_cont.append('      ,' + amend_col)
            self.output_prod_command_cont.append('FROM VP_MCIF.EVENT_BKL_REPAY_TXN')
            self.output_prod_command_cont.append("WHERE ACCT_NBR = '" + self.encrypt_acn + "'")
            self.output_prod_command_cont.append("AND JOURNAL_NBR = '" + journal_nbr + "'")
            self.output_prod_command_cont.append("AND TERM_NBR ='" + term_nbr + "'")
            self.output_prod_command_cont.append("AND BUCKET_SEQ_NBR = '" + bucket_seq_nbr + "';")
            
            # 測試使用指令
            self.output_test_command_cont.append('/*** RPYC(AMEND)部份 ***/')
            self.output_test_command_cont.append('-- EVENT_BKL_REPAY_TXN 台幣放款還款紀錄檔')
            self.output_test_command_cont.append('-------------------------------------------------------------------------------------------')
            self.output_test_command_cont.append('INSERT INTO DP_MCIF.EVENT_BKL_REPAY_TXN')
            self.output_test_command_cont.append('(')
            self.output_test_command_cont.append('INSTITUTE_ID, ACCT_NBR, TXN_DATE, JOURNAL_NBR, TERM_NBR, BUCKET_SEQ_NBR')
            amend_cols = ''
            for amend_col,before_amend,after_amend in amend_list:
                amend_cols += ',' + amend_col
            self.output_test_command_cont.append(amend_cols)
            self.output_test_command_cont.append(')')
            self.output_test_command_cont.append('VALUES')
            self.output_test_command_cont.append('(')
            self.output_test_command_cont.append("'XXX','" + acct_acn + "','" + self.data_ym[:4] + "-" + self.data_ym[4:] + "-01','" + journal_nbr + "','" + term_nbr + "','" + bucket_seq_nbr + "'")
            before_amends = ''
            for amend_col,before_amend,after_amend in amend_list:
                before_amends += ',' + before_amend
            self.output_test_command_cont.append(before_amends)
            self.output_test_command_cont.append(');')
            self.output_test_command_cont.append('')

    def gen_rpyc_delete_cont(self,work_sheet,icontect_no,key_row,key_col,acct_acn,journal_nbr,term_nbr,bucket_seq_nbr):
        # 上線程式
        self.output_result_cont.append('/*** RPYC(DELETE)部份 ***/')
        self.output_result_cont.append('-- EVENT_BKL_REPAY_TXN 台幣放款還款紀錄檔')
        self.output_result_cont.append('-------------------------------------------------------------------------------------------')
        self.output_result_cont.append("SELECT '" + icontect_no + "'||DATE||' '||TIME||' BEFORE-VALUE=>' SEQ")
        self.output_result_cont.append('      ,INSTITUTE_ID, ACCT_NBR, TXN_DATE, JOURNAL_NBR, TERM_NBR, BUCKET_SEQ_NBR')
        self.output_result_cont.append('FROM DP_MCIF.EVENT_BKL_REPAY_TXN')
        self.output_result_cont.append("WHERE ACCT_NBR = '" + acct_acn + "'")
        self.output_result_cont.append("AND JOURNAL_NBR = '" + journal_nbr + "'")
        self.output_result_cont.append("AND TERM_NBR ='" + term_nbr + "'")
        self.output_result_cont.append("AND BUCKET_SEQ_NBR = '" + bucket_seq_nbr + "';")
        self.output_result_cont.append('')
        self.output_result_cont.append('DELETE DP_MCIF.EVENT_BKL_REPAY_TXN')
        self.output_result_cont.append("WHERE ACCT_NBR = '" + acct_acn + "'")
        self.output_result_cont.append("AND JOURNAL_NBR = '" + journal_nbr + "'")
        self.output_result_cont.append("AND TERM_NBR ='" + term_nbr + "'")
        self.output_result_cont.append("AND BUCKET_SEQ_NBR = '" + bucket_seq_nbr + "';")
        self.output_result_cont.append('')
        self.output_result_cont.append("SELECT '" + icontect_no + "'||DATE||' '||TIME||' AFTER-VALUE=>' SEQ")
        self.output_result_cont.append('      ,INSTITUTE_ID, ACCT_NBR, TXN_DATE, JOURNAL_NBR, TERM_NBR, BUCKET_SEQ_NBR')
        self.output_result_cont.append('FROM DP_MCIF.EVENT_BKL_REPAY_TXN')
        self.output_result_cont.append("WHERE ACCT_NBR = '" + acct_acn + "'")
        self.output_result_cont.append("AND JOURNAL_NBR = '" + journal_nbr + "'")
        self.output_result_cont.append("AND TERM_NBR ='" + term_nbr + "'")
        self.output_result_cont.append("AND BUCKET_SEQ_NBR = '" + bucket_seq_nbr + "';")
        self.output_result_cont.append('')
        
        # 查看營運資料指令
        self.output_prod_command_cont.append('/*** RPYC(DELETE)部份 ***/')
        self.output_prod_command_cont.append('-- EVENT_BKL_REPAY_TXN 台幣放款還款紀錄檔')
        self.output_prod_command_cont.append('-------------------------------------------------------------------------------------------')
        self.output_prod_command_cont.append("SELECT '" + icontect_no + "'||DATE||' '||TIME||' BEFORE-VALUE=>' SEQ")
        self.output_prod_command_cont.append('      ,INSTITUTE_ID, ACCT_NBR, TXN_DATE, JOURNAL_NBR, TERM_NBR, BUCKET_SEQ_NBR')
        self.output_prod_command_cont.append('FROM VP_MCIF.EVENT_BKL_REPAY_TXN')
        self.output_prod_command_cont.append("WHERE ACCT_NBR = '" + self.encrypt_acn + "'")
        self.output_prod_command_cont.append("AND JOURNAL_NBR = '" + journal_nbr + "'")
        self.output_prod_command_cont.append("AND TERM_NBR ='" + term_nbr + "'")
        self.output_prod_command_cont.append("AND BUCKET_SEQ_NBR = '" + bucket_seq_nbr + "';")
        self.output_prod_command_cont.append('')
        
        # 測試使用指令
        self.output_test_command_cont.append('/*** RPYC(DELETE)部份 ***/')
        self.output_test_command_cont.append('-- EVENT_BKL_REPAY_TXN 台幣放款還款紀錄檔')
        self.output_test_command_cont.append('-------------------------------------------------------------------------------------------')
        self.output_test_command_cont.append('INSERT INTO DP_MCIF.EVENT_BKL_REPAY_TXN')
        self.output_test_command_cont.append('(INSTITUTE_ID, ACCT_NBR, TXN_DATE, JOURNAL_NBR, TERM_NBR, BUCKET_SEQ_NBR)')
        self.output_test_command_cont.append('VALUES')
        self.output_test_command_cont.append("('XXX','" + acct_acn + "','" + self.data_ym[:4] + "-" + self.data_ym[4:] + "-01','" + journal_nbr + "','" + term_nbr + "','" + bucket_seq_nbr + "');")
        self.output_test_command_cont.append('')

    def gen_boct(self,work_sheet):
        keys_coordinate = get_keys_coordinate(work_sheet)
        self.output_result_cont.append('/*** BOCT部份 ***/')
        self.output_result_cont.append('-- UPDATE value of BOCT_01,20,21,25,36,46,47,48,59')
        self.output_result_cont.append('-- Col: DELETE_IND, TXN_AMT, BAL_AMT')
        self.output_result_cont.append('')
        self.output_prod_command_cont.append('/*** BOCT部份 ***/')
        self.output_prod_command_cont.append('')
        self.output_test_command_cont.append('/*** BOCT部份 ***/')
        self.output_test_command_cont.append('')
        for row, col in keys_coordinate:
            action_row, action_col = get_action_coordinate(work_sheet, row, col, search_range=2)
            action = work_sheet.cell(action_row,action_col+1).value
            acct_acn = work_sheet.cell(row,col+1).value
            record_code = work_sheet.cell(row,col+2).value
            boct_type = work_sheet.cell(row,col+3).value
            print(action,acct_acn,record_code,boct_type)
            
            if action == 'A(AMEND)' and acct_acn is not None:
                self.gen_boct_amend_cont(work_sheet,self.icontect_no,row,col,acct_acn,record_code,boct_type,self.data_ym)
            elif action == 'A(AMEND)' and acct_acn is None:
                self.output_result_cont.append('/* 無BOCT更改需求 */')
            else:
                self.log_display.error('改檔EXCEL BOCT工作表Action非"A(AMEND)"，請修正!')
        self.output_result_cont.append('')
        self.output_result_cont.append('-- No Type: 26, 40, 50, 63, 66')
        self.output_result_cont.append('')
        
    def gen_boct_amend_cont(self,work_sheet,icontect_no,key_row,key_col,acct_acn,record_code,boct_type,data_ym):
        field_name_row, field_name_col = get_field_name_coordinate(work_sheet,key_row,key_col,search_range=3)
        if field_name_row is None or field_name_col is None:
            self.log_display.error('改檔EXCEL BOCT工作表"key:"的右下格子非"Field Name"，請修正!')
        else:
            field_dict = {'01-AMOUNT':'TXN_AMT','01-BALANCE':'BAL_AMT','01-ARREARS':'ARREARS','DELI':'DELETE_IND'}
            annotation_dict = {'01':'-- EVENT_BKL_FIN_TXN  BOCT_01-帳務交易',
                               '20':'-- EVENT_BKL_REPAY_RSN_TXN  BOCT_20-還款原因',
                               '21':'-- EVENT_BKL_FEE_TXN  BOCT_21-手續費收取',
                               '25':'-- EVENT_BKL_REPAYMENT_TXN  BOCT_25-還款金額合計明細',
                               '36':'-- EVENT_BKL_CHK_REPAY_TXN  BOCT_36-票據還款',
                               '46':'-- EVENT_BKL_CHARACTER_TXN  BOCT_46-維護性質別／子類別',
                               '47':'-- EVENT_BKL_PRD_BRANCH_TXN  BOCT_47-維護產品別／分行別',
                               '48':'-- EVENT_BKL_STATUS_CHG_TXN  BOCT_48-更改帳戶狀態',
                               '59':'-- EVENT_BKL_SUBBSIDY_TXN  BOCT_59-補貼息明細'}
            table_name_dict = {'01':'EVENT_BKL_FIN_TXN',
                               '20':'EVENT_BKL_REPAY_RSN_TXN',
                               '21':'EVENT_BKL_FEE_TXN',
                               '25':'EVENT_BKL_REPAYMENT_TXN',
                               '36':'EVENT_BKL_CHK_REPAY_TXN',
                               '46':'EVENT_BKL_CHARACTER_TXN',
                               '47':'EVENT_BKL_PRD_BRANCH_TXN',
                               '48':'EVENT_BKL_STATUS_CHG_TXN',
                               '59':'EVENT_BKL_SUBBSIDY_TXN'}
            i = 1
            amend_list = []
            extract_field_list = ['01-AMOUNT','01-BALANCE','01-ARREARS']
            while work_sheet.cell(field_name_row+i,field_name_col).value is not None:
                field_name = work_sheet.cell(field_name_row+i,field_name_col).value.strip()
                # occurs = work_sheet.cell(field_name_row+i,field_name_col+1).value.strip()
                before_amend = work_sheet.cell(field_name_row+i,field_name_col+2).value.strip()
                after_amend = work_sheet.cell(field_name_row+i,field_name_col+3).value.strip()
                if field_name in extract_field_list:
                    before_amend = number_extract(before_amend)
                    after_amend = number_extract(after_amend)
                amend_col = field_dict[field_name]
                amend_list.append([amend_col,before_amend,after_amend])
                i += 1
            
            if boct_type in annotation_dict:
                # 上線程式
                self.output_result_cont.append('-------------------------------------------------------------------------------------------')
                self.output_result_cont.append(annotation_dict[boct_type])
                self.output_result_cont.append("SELECT '" + icontect_no + "'||DATE||' '||TIME||' BEFORE-VALUE=>' SEQ")
                self.output_result_cont.append('      , TXN_TYPE_CODE, POSTING_DATE, TXN_DATE, INSTITUTE_ID, ACCT_NBR, RECORD_CODE, TXN_SEQ_NBR')
                for amend_col,before_amend,after_amend in amend_list:
                    self.output_result_cont.append('      , ' + amend_col)
                self.output_result_cont.append('FROM DP_MCIF_HISTORY.' + table_name_dict[boct_type] + data_ym)
                self.output_result_cont.append("WHERE ACCT_NBR = '" + acct_acn + "'")
                self.output_result_cont.append("AND RECORD_CODE = '" + record_code + "';")
                self.output_result_cont.append('')
                self.output_result_cont.append('UPDATE DP_MCIF_HISTORY.' + table_name_dict[boct_type] + data_ym)
                self.output_result_cont.append('SET')
                for i in range(len(amend_list)):
                    comma = ' ' if i == 0 else ','
                    amend_col = amend_list[i][0]
                    after_amend = amend_list[i][2]
                    if amend_col == 'DELETE_IND':
                        self.output_result_cont.append('    ' + comma + amend_col + " = '" + after_amend + "'")
                    else:
                        self.output_result_cont.append('    ' + comma + amend_col + ' = ' + after_amend)
                self.output_result_cont.append("WHERE ACCT_NBR = '" + acct_acn + "'")
                self.output_result_cont.append("AND RECORD_CODE = '" + record_code + "';")
                self.output_result_cont.append('')
                self.output_result_cont.append("SELECT '" + icontect_no + "'||DATE||' '||TIME||' AFTER-VALUE=>' SEQ")
                self.output_result_cont.append('      , TXN_TYPE_CODE, POSTING_DATE, TXN_DATE, INSTITUTE_ID, ACCT_NBR, RECORD_CODE, TXN_SEQ_NBR')
                for amend_col,before_amend,after_amend in amend_list:
                    self.output_result_cont.append('      , ' + amend_col)
                self.output_result_cont.append('FROM DP_MCIF_HISTORY.' + table_name_dict[boct_type] + data_ym)
                self.output_result_cont.append("WHERE ACCT_NBR = '" + acct_acn + "'")
                self.output_result_cont.append("AND RECORD_CODE = '" + record_code + "';")
                self.output_result_cont.append('')
                
                # 查看營運資料指令
                self.output_prod_command_cont.append('-------------------------------------------------------------------------------------------')
                self.output_prod_command_cont.append(annotation_dict[boct_type])
                self.output_prod_command_cont.append("SELECT '" + icontect_no + "'||DATE||' '||TIME||' BEFORE-VALUE=>' SEQ")
                self.output_prod_command_cont.append('      , TXN_TYPE_CODE, POSTING_DATE, TXN_DATE, INSTITUTE_ID, ACCT_NBR, RECORD_CODE, TXN_SEQ_NBR')
                for amend_col,before_amend,after_amend in amend_list:
                    self.output_prod_command_cont.append('      , ' + amend_col)
                self.output_prod_command_cont.append('FROM VP_MCIF.' + table_name_dict[boct_type] + data_ym)
                self.output_prod_command_cont.append("WHERE ACCT_NBR = '" + self.encrypt_acn + "'")
                self.output_prod_command_cont.append("AND RECORD_CODE = '" + record_code + "';")
                self.output_prod_command_cont.append('')
                
                # 測試使用指令
                self.output_test_command_cont.append('-------------------------------------------------------------------------------------------')
                self.output_test_command_cont.append(annotation_dict[boct_type])
                self.output_test_command_cont.append('--DROP TABLE DP_MCIF_HISTORY.' + table_name_dict[boct_type] + data_ym + ';')
                self.output_test_command_cont.append('--CREATE TABLE DP_MCIF_HISTORY.' + table_name_dict[boct_type] + data_ym + ' AS (SELECT * FROM DP_MCIF.' + table_name_dict[boct_type] + ') WITH NO DATA;')
                self.output_test_command_cont.append('INSERT INTO DP_MCIF_HISTORY.' + table_name_dict[boct_type] + data_ym)
                self.output_test_command_cont.append('(')
                self.output_test_command_cont.append('TXN_TYPE_CODE, POSTING_DATE, TXN_DATE, INSTITUTE_ID, ACCT_NBR, RECORD_CODE, TXN_SEQ_NBR')
                for amend_col,before_amend,after_amend in amend_list:
                    self.output_test_command_cont.append(', ' + amend_col)
                self.output_test_command_cont.append(')')
                self.output_test_command_cont.append('VALUES')
                self.output_test_command_cont.append('(')
                self.output_test_command_cont.append("'" + boct_type + "','" + self.data_ym[:4] + "-" + self.data_ym[4:] + "-01','" + self.data_ym[:4] + "-" + self.data_ym[4:] + "-01','XXX','" + acct_acn + "','" + record_code + "','XXXXXXX'")
                for amend_col,before_amend,after_amend in amend_list:
                    self.output_test_command_cont.append(", '" + before_amend + "'")
                self.output_test_command_cont.append(');')
                self.output_test_command_cont.append('')
    
    def gen_boct_q(self,work_sheet):
        keys_coordinate = get_keys_coordinate(work_sheet)
        self.output_result_cont.append('/*** BOCT_Q部份 ***/')
        self.output_result_cont.append('-- UPDATE value of BOCT_01,20,21,25,36,46,47,48,59')
        self.output_result_cont.append('-- Col: DELETE_IND, TXN_AMT, BAL_AMT')
        self.output_result_cont.append('')
        self.output_prod_command_cont.append('/*** BOCT_Q部份 ***/')
        self.output_prod_command_cont.append('')
        self.output_test_command_cont.append('/*** BOCT_Q部份 ***/')
        self.output_test_command_cont.append('')
        for row, col in keys_coordinate:
            action_row, action_col = get_action_coordinate(work_sheet, row, col, search_range=2)
            action = work_sheet.cell(action_row,action_col+1).value.upper()
            acct_acn = work_sheet.cell(row,col+1).value
            record_code = work_sheet.cell(row,col+2).value
            txn_seq_nbr = work_sheet.cell(row,col+3).value
            print(action,acct_acn,record_code,txn_seq_nbr)
            
            if action == 'A(AMEND)' and acct_acn is not None:
                self.gen_boct_q_amend_cont(work_sheet,self.icontect_no,row,col,acct_acn,txn_seq_nbr,self.data_ym)
            elif action == 'A(AMEND)' and acct_acn is None:
                self.output_result_cont.append('/* 無BOCT_Q更改需求 */')
            else:
                self.log_display.error('改檔EXCEL BOCT_Q工作表Action非"A(AMEND)"，請修正!')
                return
        self.output_result_cont.append('')
        self.output_result_cont.append('-- No Type: 26, 40, 50, 63, 66')
        self.output_result_cont.append('')

    def gen_boct_q_amend_cont(self,work_sheet,icontect_no,key_row,key_col,acct_acn,txn_seq_nbr,data_ym):
        field_name_row, field_name_col = get_field_name_coordinate(work_sheet,key_row,key_col,search_range=3)
        if field_name_row is None or field_name_col is None:
            self.log_display.error('改檔EXCEL BOCT工作表"key:"的右下格子非"Field Name"，請修正!')
            return
        else:
            field_dict = {'01-AMOUNT':'TXN_AMT','01-BALANCE':'BAL_AMT','01-ARREARS':'ARREARS','DELI':'DELETE_IND'}
            annotation_dict = {'01':'-- EVENT_BKL_FIN_TXN  BOCT_01-帳務交易',
                               '20':'-- EVENT_BKL_REPAY_RSN_TXN  BOCT_20-還款原因',
                               '21':'-- EVENT_BKL_FEE_TXN  BOCT_21-手續費收取',
                               '25':'-- EVENT_BKL_REPAYMENT_TXN  BOCT_25-還款金額合計明細',
                               '36':'-- EVENT_BKL_CHK_REPAY_TXN  BOCT_36-票據還款',
                               '46':'-- EVENT_BKL_CHARACTER_TXN  BOCT_46-維護性質別／子類別',
                               '47':'-- EVENT_BKL_PRD_BRANCH_TXN  BOCT_47-維護產品別／分行別',
                               '48':'-- EVENT_BKL_STATUS_CHG_TXN  BOCT_48-更改帳戶狀態',
                               '59':'-- EVENT_BKL_SUBBSIDY_TXN  BOCT_59-補貼息明細'}
            table_name_dict = {'01':'EVENT_BKL_FIN_TXN',
                               '20':'EVENT_BKL_REPAY_RSN_TXN',
                               '21':'EVENT_BKL_FEE_TXN',
                               '25':'EVENT_BKL_REPAYMENT_TXN',
                               '36':'EVENT_BKL_CHK_REPAY_TXN',
                               '46':'EVENT_BKL_CHARACTER_TXN',
                               '47':'EVENT_BKL_PRD_BRANCH_TXN',
                               '48':'EVENT_BKL_STATUS_CHG_TXN',
                               '59':'EVENT_BKL_SUBBSIDY_TXN'}
            i = 1
            amend_list = []
            extract_field_list = ['01-AMOUNT','01-BALANCE','01-ARREARS']
            while work_sheet.cell(field_name_row+i,field_name_col).value is not None:
                field_name = work_sheet.cell(field_name_row+i,field_name_col).value.strip()
                boct_type = work_sheet.cell(field_name_row+i,field_name_col+1).value.strip()
                rec_no = work_sheet.cell(field_name_row+i,field_name_col+2).value.strip()
                after_amend = work_sheet.cell(field_name_row+i,field_name_col+3).value.strip()
                if field_name in extract_field_list:
                    # before_amend = number_extract(before_amend)
                    after_amend = number_extract(after_amend)
                amend_col = field_dict[field_name]
                amend_list.append([amend_col,boct_type,rec_no,after_amend])
                i += 1
                
            for amend_col,boct_type,rec_no,after_amend in amend_list:
                if boct_type in annotation_dict:
                    # 上線程式
                    self.output_result_cont.append('-------------------------------------------------------------------------------------------')
                    self.output_result_cont.append(annotation_dict[boct_type])
                    self.output_result_cont.append("SELECT '" + icontect_no + "'||DATE||' '||TIME||' BEFORE-VALUE=>' SEQ")
                    self.output_result_cont.append('      , TXN_TYPE_CODE, POSTING_DATE, TXN_DATE, INSTITUTE_ID, ACCT_NBR, RECORD_CODE, TXN_SEQ_NBR')
                    self.output_result_cont.append('      , ' + amend_col)
                    self.output_result_cont.append('FROM DP_MCIF_HISTORY.' + table_name_dict[boct_type] + data_ym)
                    self.output_result_cont.append("WHERE ACCT_NBR = '" + acct_acn + "'")
                    self.output_result_cont.append("AND RECORD_CODE = '" + rec_no + "';")
                    self.output_result_cont.append('')
                    self.output_result_cont.append('UPDATE DP_MCIF_HISTORY.' + table_name_dict[boct_type] + data_ym)
                    if amend_col == 'DELETE_IND':
                        self.output_result_cont.append('SET ' + amend_col + " = '" + after_amend + "'")
                    else:
                        self.output_result_cont.append('SET ' + amend_col + ' = ' + after_amend)
                    self.output_result_cont.append("WHERE ACCT_NBR = '" + acct_acn + "'")
                    self.output_result_cont.append("AND RECORD_CODE = '" + rec_no + "';")
                    self.output_result_cont.append('')
                    self.output_result_cont.append("SELECT '" + icontect_no + "'||DATE||' '||TIME||' AFTER-VALUE=>' SEQ")
                    self.output_result_cont.append('      , TXN_TYPE_CODE, POSTING_DATE, TXN_DATE, INSTITUTE_ID, ACCT_NBR, RECORD_CODE, TXN_SEQ_NBR')
                    self.output_result_cont.append('      , ' + amend_col)
                    self.output_result_cont.append('FROM DP_MCIF_HISTORY.' + table_name_dict[boct_type] + data_ym)
                    self.output_result_cont.append("WHERE ACCT_NBR = '" + acct_acn + "'")
                    self.output_result_cont.append("AND RECORD_CODE = '" + rec_no + "';")
                    
                    # 查看營運資料指令
                    self.output_prod_command_cont.append('-------------------------------------------------------------------------------------------')
                    self.output_prod_command_cont.append(annotation_dict[boct_type])
                    self.output_prod_command_cont.append("SELECT '" + icontect_no + "'||DATE||' '||TIME||' BEFORE-VALUE=>' SEQ")
                    self.output_prod_command_cont.append('      , TXN_TYPE_CODE, POSTING_DATE, TXN_DATE, INSTITUTE_ID, ACCT_NBR, RECORD_CODE, TXN_SEQ_NBR')
                    self.output_prod_command_cont.append('      , ' + amend_col)
                    self.output_prod_command_cont.append('FROM VP_MCIF.' + table_name_dict[boct_type] + data_ym)
                    self.output_prod_command_cont.append("WHERE ACCT_NBR = '" + self.encrypt_acn + "'")
                    self.output_prod_command_cont.append("AND RECORD_CODE = '" + rec_no + "';")
                    self.output_prod_command_cont.append('')
                    
                    # 測試使用指令
                    self.output_test_command_cont.append('-------------------------------------------------------------------------------------------')
                    self.output_test_command_cont.append(annotation_dict[boct_type])
                    self.output_test_command_cont.append('--DROP TABLE DP_MCIF_HISTORY.' + table_name_dict[boct_type] + data_ym + ';')
                    self.output_test_command_cont.append('--CREATE TABLE DP_MCIF_HISTORY.' + table_name_dict[boct_type] + data_ym + ' AS (SELECT * FROM DP_MCIF.' + table_name_dict[boct_type] + ') WITH NO DATA;')
                    self.output_test_command_cont.append('INSERT INTO DP_MCIF_HISTORY.' + table_name_dict[boct_type] + data_ym)
                    self.output_test_command_cont.append('(')
                    self.output_test_command_cont.append('TXN_TYPE_CODE, POSTING_DATE, TXN_DATE, INSTITUTE_ID, ACCT_NBR, RECORD_CODE, TXN_SEQ_NBR')
                    self.output_test_command_cont.append(', ' + amend_col)
                    self.output_test_command_cont.append(')')
                    self.output_test_command_cont.append('VALUES')
                    self.output_test_command_cont.append('(')
                    self.output_test_command_cont.append("'" + boct_type + "','" + self.data_ym[:4] + "-" + self.data_ym[4:] + "-01','" + self.data_ym[:4] + "-" + self.data_ym[4:] + "-01','XXX','" + acct_acn + "','" + rec_no + "','" + txn_seq_nbr + "'")
                    if amend_col == 'DELETE_IND':
                        self.output_test_command_cont.append(", '0'")
                    else:
                        self.output_test_command_cont.append(", 0")
                    self.output_test_command_cont.append(");")
                    self.output_test_command_cont.append('')

